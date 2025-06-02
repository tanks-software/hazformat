import streamlit as st
import json
import tempfile
import pandas as pd
import textwrap
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io
import os
import time

# ----------------- Google API Helper ------------------

SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/spreadsheets"
]

class GoogleDriveSheets:
    def __init__(self, service_account_info):
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".json")
        tmp.write(json.dumps(service_account_info).encode())
        tmp.close()
        self.SERVICE_ACCOUNT_FILE = tmp.name

        self.creds = service_account.Credentials.from_service_account_file(
            self.SERVICE_ACCOUNT_FILE, scopes=SCOPES
        )
        self.drive_service = build('drive', 'v3', credentials=self.creds)
        self.sheets_service = build('sheets', 'v4', credentials=self.creds)

    def download_spreadsheet_as_df(self, spreadsheet_id, sheet_name):
        try:
            result = self.sheets_service.spreadsheets().values().get(
                spreadsheetId=spreadsheet_id,
                range=sheet_name
            ).execute()
            values = result.get('values', [])
            if not values:
                return pd.DataFrame()

            if sheet_name == "Equipment Type":
                df = pd.DataFrame(values, columns=["Equipment Type"])
            else:
                header = values[0]
                data_rows = values[1:]
                max_cols = len(header)
                padded_rows = []
                for row in data_rows:
                    if len(row) < max_cols:
                        row += [""] * (max_cols - len(row))
                    elif len(row) > max_cols:
                        row = row[:max_cols]
                    padded_rows.append(row)
                df = pd.DataFrame(padded_rows, columns=header)
            #st.write(f"Debug: Fetched {len(df)} rows from sheet '{sheet_name}'")
            return df
        except Exception as e:
            st.error(f"Error fetching data from sheet '{sheet_name}': {str(e)}")
            return pd.DataFrame()

    def append_to_sheet(self, spreadsheet_id, sheet_name, values):
        try:
            body = {'values': [values]}
            result = self.sheets_service.spreadsheets().values().append(
                spreadsheetId=spreadsheet_id,
                range=sheet_name,
                valueInputOption="RAW",
                body=body
            ).execute()
            return result
        except Exception as e:
            st.error(f"Error appending to sheet '{sheet_name}': {str(e)}")
            return None

    def list_folder_files(self, folder_id, mime_types=None):
        query = f"'{folder_id}' in parents and trashed = false"
        if mime_types:
            mime_filter = " or ".join([f"mimeType='{mime}'" for mime in mime_types])
            query += f" and ({mime_filter})"
        files = []
        page_token = None
        while True:
            response = self.drive_service.files().list(
                q=query,
                spaces='drive',
                fields='nextPageToken, files(id, name, mimeType)',
                pageToken=page_token
            ).execute()
            files.extend(response.get('files', []))
            page_token = response.get('nextPageToken', None)
            if page_token is None:
                break
        return files

    def download_file_to_temp(self, file_id, file_name):
        request = self.drive_service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        fh.seek(0)
        suffix = os.path.splitext(file_name)[1]
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        tmp.write(fh.read())
        tmp.flush()
        tmp.close()
        return tmp.name

# ----------------- Helper functions ------------------

def convert_keys_to_template_style(orig_dict):
    key_map = {
        "technicalName": "TECHNICAL_NAME",
        "class": "CLASS",
        "unno": "UNNO",
        "subrisk": "SUBRISK",
        "packingGroup": "PACKING_GROUP",
        "ems": "EMS",
        "flashPoint": "FLASH_POINT",
        "marinePollutant": "MARINE_POLLUTANT",
        "Limited Quantity ": "LIMITED_QUANTITY",
        "natureOfCargo": "NATURE_OF_CARGO",
        "MFAG Number": "MFAG_NUMBER"
    }
    return {key_map.get(k, k).replace(" ", "_"): v for k, v in orig_dict.items()}

def format_address_by_chars(address, width=40):
    if not address:
        return ""
    return "\n".join(textwrap.wrap(address, width=width))

def fill_excel_template(template_path, data):
    wb = load_workbook(template_path)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for key, val in data.items():
                        placeholder = "{{" + key + "}}"
                        if placeholder in cell.value:
                            if key in ("SHIPPER_ADDRESS", "CONSIGNEE_ADDRESS"):
                                val = val.replace("\n", "\n")
                                cell.value = cell.value.replace(placeholder, val)
                                cell.alignment = Alignment(wrap_text=True)
                            else:
                                cell.value = cell.value.replace(placeholder, str(val))
    tmp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp_xlsx.name)
    return tmp_xlsx.name

# ----------------- Load Master Data ------------------

@st.cache_data(ttl=3600, show_spinner=False)
def load_master_data_from_drive(service_account_info, spreadsheet_id, _cache_buster):
    gds = GoogleDriveSheets(service_account_info)
    cargo_df = gds.download_spreadsheet_as_df(spreadsheet_id, "cargo")
    equipment_df = gds.download_spreadsheet_as_df(spreadsheet_id, "Equipment Type")
    shippers_df = gds.download_spreadsheet_as_df(spreadsheet_id, "Shippers")
    consignees_df = gds.download_spreadsheet_as_df(spreadsheet_id, "Consignees")
    ports_df = gds.download_spreadsheet_as_df(spreadsheet_id, "Ports")

    vessels_df = pd.DataFrame()
    try:
        vessels_df = gds.download_spreadsheet_as_df(spreadsheet_id, "Vessels")
    except Exception:
        vessels_df = pd.DataFrame()

    cargo_data = {}
    for _, row in cargo_df.iterrows():
        cargo_name = row["Proper Shipping Name"]
        details = row.drop(labels=["Proper Shipping Name"]).to_dict()
        details = {k: ("" if pd.isna(v) else v) for k, v in details.items()}
        cargo_data[cargo_name] = details

    equipment_types = equipment_df.iloc[:, 0].dropna().tolist()
    shippers = shippers_df["Shipper"].dropna().tolist()
    shipper_contacts = shippers_df.set_index("Shipper")[["ContactName", "ContactNumber", "Shipper_Address"]].to_dict(orient="index")
    consignees = consignees_df["Consignee"].dropna().tolist()
    consignee_addresses = consignees_df.set_index("Consignee")["Consignee_Address"].to_dict()

    pol_ports = ports_df["POL"].dropna().tolist() if "POL" in ports_df.columns else []
    pod_ports = ports_df["POD"].dropna().tolist() if "POD" in ports_df.columns else []

    vessels = vessels_df["Vessel_Name"].dropna().tolist() if not vessels_df.empty else []

    #st.write(f"Debug: Shippers list: {shippers}")
    return cargo_data, shippers, shipper_contacts, consignees, consignee_addresses, pol_ports, pod_ports, equipment_types, vessels

# ----------------- Load Templates ------------------

@st.cache_resource(ttl=3600, show_spinner=False)
def load_templates_from_drive(service_account_info, templates_folder_id):
    gds = GoogleDriveSheets(service_account_info)
    files = gds.list_folder_files(templates_folder_id, mime_types=[
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ])
    templates = {}
    for f in files:
        path = gds.download_file_to_temp(f["id"], f["name"])
        name = os.path.splitext(f["name"])[0]
        templates[name] = path
    return templates

# ----------------- Main App ------------------

def main():
    service_account_json_str = st.secrets["google"]["service_account_json"]
    service_account_info = json.loads(service_account_json_str)

    TEMPLATES_FOLDER_ID = "1cYa1GvGGtKxP1TPD-ACzA281LpxWZNmy"
    MASTER_SPREADSHEET_ID = "1Hj0XV7Pe4oC7Mu7_x7jNvuJzOmDT5qZtlhqSytcdDno"

    # Initialize session state for cache busting and form visibility
    if 'cache_buster' not in st.session_state:
        st.session_state.cache_buster = int(time.time())
    if 'show_shipper_form' not in st.session_state:
        st.session_state.show_shipper_form = False
    if 'show_consignee_form' not in st.session_state:
        st.session_state.show_consignee_form = False
    if 'show_pod_form' not in st.session_state:
        st.session_state.show_pod_form = False
    if 'show_vessel_form' not in st.session_state:
        st.session_state.show_vessel_form = False

    # Load master data
    cargo_data_raw, shippers, shipper_contacts, consignees, consignee_addresses, pol_ports, pod_ports, equipment_types, vessels = load_master_data_from_drive(
        service_account_info, MASTER_SPREADSHEET_ID, st.session_state.cache_buster
    )
    cargo_data = {k: convert_keys_to_template_style(v) for k, v in cargo_data_raw.items()}

    templates = load_templates_from_drive(service_account_info, TEMPLATES_FOLDER_ID)

    st.title("Hazardous DG Form Generator")

    # Template selection
    template_options = ["Select a Template"] + list(templates.keys())
    selected_template_name = st.selectbox("Select Template", template_options)

    if selected_template_name == "Select a Template":
        st.warning("Please select a template to proceed.")
        st.stop()
    else:
        selected_template_path = templates[selected_template_name]

    # Form inputs
    st.subheader("Cargo Information")
    cargo = st.selectbox("Select Proper Shipping Name", list(cargo_data.keys()))
    if cargo and cargo in cargo_data:
        details = cargo_data[cargo]
    else:
        details = {k: "" for k in [
            "TECHNICAL_NAME", "CLASS", "UNNO", "SUBRISK", "PACKING_GROUP", "EMS",
            "FLASH_POINT", "MARINE_POLLUTANT", "LIMITED_QUANTITY", "NATURE_OF_CARGO", "MFAG_NUMBER"]}

    technical_name = st.text_input("Technical Name", details.get("TECHNICAL_NAME", ""))
    cargo_class = st.text_input("Class", details.get("CLASS", ""))
    unno = st.text_input("UNNO", details.get("UNNO", ""))
    subrisk = st.text_input("Subrisk", details.get("SUBRISK", ""))
    packing_group = st.text_input("Packing Group", details.get("PACKING_GROUP", ""))
    ems = st.text_input("EMS", details.get("EMS", ""))
    flash_point = st.text_input("Flash Point", details.get("FLASH_POINT", ""))

    marine_pollutant_options = ["YES", "NO"]
    marine_pollutant_default = details.get("MARINE_POLLUTANT", "").upper()
    marine_pollutant = st.selectbox(
        "Marine Pollutant", marine_pollutant_options,
        index=marine_pollutant_options.index(marine_pollutant_default) if marine_pollutant_default in marine_pollutant_options else 1)

    limited_quantity_options = ["YES", "NO", "-"]
    limited_quantity_default = details.get("LIMITED_QUANTITY", "").upper()
    limited_quantity = st.selectbox(
        "Limited Quantity", limited_quantity_options,
        index=limited_quantity_options.index(limited_quantity_default) if limited_quantity_default in limited_quantity_options else 2)

    nature_of_cargo_options = ["SOLID", "LIQUID", "GAS", "-"]
    nature_of_cargo_default = details.get("NATURE_OF_CARGO", "").upper()
    nature_of_cargo = st.selectbox(
        "Nature of Cargo", nature_of_cargo_options,
        index=nature_of_cargo_options.index(nature_of_cargo_default) if nature_of_cargo_default in nature_of_cargo_options else 3)

    mfag_number = st.text_input("MFAG Number", details.get("MFAG_NUMBER", ""))

    # Shipper selection with "Add a Shipper" at the top, second option as default
    st.subheader("Shipper Details")
    shipper_options = ["Add a Shipper"] + shippers
    default_shipper_index = 1 if len(shipper_options) > 1 else 0
    shipper = st.selectbox("Shipper", shipper_options, index=default_shipper_index, key=f"shipper_select_{st.session_state.cache_buster}")
    
    if shipper == "Add a Shipper":
        st.session_state.show_shipper_form = True
    elif shipper in shippers:
        st.session_state.show_shipper_form = False

    # Form for adding new Shipper
    if st.session_state.show_shipper_form:
        with st.expander("Add New Shipper", expanded=True):
            new_shipper = st.text_input("New Shipper Name", key="new_shipper")
            new_shipper_contact = st.text_input("New Shipper Contact Name", key="new_shipper_contact")
            new_shipper_number = st.text_input("New Shipper Contact Number", key="new_shipper_number")
            new_shipper_address = st.text_area("New Shipper Address", key="new_shipper_address")
            if st.button("Add Shipper"):
                if new_shipper.strip():
                    gds = GoogleDriveSheets(service_account_info)
                    values = [new_shipper.strip(), new_shipper_contact.strip(), new_shipper_number.strip(), new_shipper_address.strip()]
                    result = gds.append_to_sheet(MASTER_SPREADSHEET_ID, "Shippers", values)
                    if result:
                        load_master_data_from_drive.clear()
                        st.session_state.cache_buster = int(time.time())
                        st.session_state.show_shipper_form = False
                        st.success(f"Shipper '{new_shipper}' added successfully!")
                        time.sleep(2)
                        st.rerun()
                    else:
                        st.error("Failed to add shipper. Please try again.")
                else:
                    st.error("Shipper name is required.")

    if shipper and shipper in shipper_contacts:
        contact_name = shipper_contacts[shipper]["ContactName"]
        contact_number = shipper_contacts[shipper]["ContactNumber"]
        shipper_address = shipper_contacts[shipper].get("Shipper_Address", "")
    else:
        contact_name = ""
        contact_number = ""
        shipper_address = ""

    formatted_shipper_address = format_address_by_chars(shipper_address, width=40)
    shipper_contact_name = st.text_input("Shipper Contact Name", contact_name)
    shipper_contact_number = st.text_input("Shipper Contact Number", contact_number)
    shipper_address_text = st.text_area("Shipper Address", formatted_shipper_address)

    # Consignee selection with "Add a Consignee" at the top, second option as default
    st.subheader("Consignee Details")
    consignee_options = ["Add a Consignee"] + consignees
    default_consignee_index = 1 if len(consignee_options) > 1 else 0
    consignee = st.selectbox("Consignee", consignee_options, index=default_consignee_index, key=f"consignee_select_{st.session_state.cache_buster}")
    
    if consignee == "Add a Consignee":
        st.session_state.show_consignee_form = True
    elif consignee in consignees:
        st.session_state.show_consignee_form = False

    # Form for adding new Consignee
    if st.session_state.show_consignee_form:
        with st.expander("Add New Consignee", expanded=True):
            new_consignee = st.text_input("New Consignee Name", key="new_consignee")
            new_consignee_address = st.text_area("New Consignee Address", key="new_consignee_address")
            if st.button("Add Consignee"):
                if new_consignee.strip():
                    gds = GoogleDriveSheets(service_account_info)
                    values = [new_consignee.strip(), new_consignee_address.strip()]
                    result = gds.append_to_sheet(MASTER_SPREADSHEET_ID, "Consignees", values)
                    if result:
                        load_master_data_from_drive.clear()
                        st.session_state.cache_buster = int(time.time())
                        st.session_state.show_consignee_form = False
                        st.success(f"Consignee '{new_consignee}' added successfully!")
                        time.sleep(2)
                        st.rerun()
                    else:
                        st.error("Failed to add consignee. Please try again.")
                else:
                    st.error("Consignee name is required.")

    if consignee and consignee in consignee_addresses:
        consignee_address = consignee_addresses[consignee]
    else:
        consignee_address = ""

    formatted_consignee_address = format_address_by_chars(consignee_address, width=40)
    consignee_address_text = st.text_area("Consignee Address", formatted_consignee_address)

    # Port of Loading (POL) without "Add a POL"
    st.subheader("Port Details")
    pol = st.selectbox("Port of Loading (POL)", pol_ports, key=f"pol_select_{st.session_state.cache_buster}")

    # Port of Discharge (POD) with "Add a POD" at the top, second option as default
    pod_options = ["Add a POD"] + pod_ports
    default_pod_index = 1 if len(pod_options) > 1 else 0
    pod = st.selectbox("Port of Discharge (POD)", pod_options, index=default_pod_index, key=f"pod_select_{st.session_state.cache_buster}")
    
    if pod == "Add a POD":
        st.session_state.show_pod_form = True
    elif pod in pod_ports:
        st.session_state.show_pod_form = False

    # Form for adding new POD
    if st.session_state.show_pod_form:
        with st.expander("Add New POD", expanded=True):
            new_pod = st.text_input("New Port of Discharge (POD)", key="new_pod")
            if st.button("Add POD"):
                if new_pod.strip():
                    gds = GoogleDriveSheets(service_account_info)
                    values = ["", new_pod.strip()]
                    result = gds.append_to_sheet(MASTER_SPREADSHEET_ID, "Ports", values)
                    if result:
                        load_master_data_from_drive.clear()
                        st.session_state.cache_buster = int(time.time())
                        st.session_state.show_pod_form = False
                        st.success(f"POD '{new_pod}' added successfully!")
                        time.sleep(2)
                        st.rerun()
                    else:
                        st.error("Failed to add POD. Please try again.")
                else:
                    st.error("POD name is required.")

    # Vessel with "Add a Vessel" at the top, second option as default
    vessel_options = ["Add a Vessel"] + vessels
    default_vessel_index = 1 if len(vessel_options) > 1 else 0
    vessel = st.selectbox("Vessel", vessel_options, index=default_vessel_index, key=f"vessel_select_{st.session_state.cache_buster}")
    
    if vessel == "Add a Vessel":
        st.session_state.show_vessel_form = True
    elif vessel in vessels:
        st.session_state.show_vessel_form = False

    # Form for adding new Vessel
    if st.session_state.show_vessel_form:
        with st.expander("Add New Vessel", expanded=True):
            new_vessel = st.text_input("New Vessel Name", key="new_vessel")
            if st.button("Add Vessel"):
                if new_vessel.strip():
                    gds = GoogleDriveSheets(service_account_info)
                    values = [new_vessel.strip()]
                    result = gds.append_to_sheet(MASTER_SPREADSHEET_ID, "Vessels", values)
                    if result:
                        load_master_data_from_drive.clear()
                        st.session_state.cache_buster = int(time.time())
                        st.session_state.show_vessel_form = False
                        st.success(f"Vessel '{new_vessel}' added successfully!")
                        time.sleep(2)
                        st.rerun()
                    else:
                        st.error("Failed to add vessel. Please try again.")
                else:
                    st.error("Vessel name is required.")

    # Equipment Type (no "Add a Equipment Type" option)
    st.subheader("Cargo Details")
    qty = st.number_input("Quantity", min_value=1, step=1, value=1)
    equipment_type = st.selectbox("Equipment Type *", equipment_types, key=f"equipment_select_{st.session_state.cache_buster}")

    outer_package = st.text_input("Outer Package *")
    inner_package = st.text_input("Inner Package *")
    gross_wt = st.text_input("Gross Weight *")
    net_wt = st.text_input("Net Weight *")
    container_number = st.text_input("Container Number (Optional)")
    seal_number = st.text_input("Seal Number (Optional)")

    mandatory_fields_filled = all([
        outer_package.strip(),
        inner_package.strip(),
        gross_wt.strip(),
        net_wt.strip(),
        equipment_type != ""
    ])

    combined_qty_equipment = f"{qty}x{equipment_type}" if equipment_type != "" and qty > 0 else ""

    if st.button("Generate Document"):
        if not mandatory_fields_filled:
            st.error("Please fill all mandatory fields marked with * and select Equipment Type")
        else:
            data = {
                "SHIPPER": shipper if shipper != "Add a Shipper" else "",
                "SHIPPER_CONTACT_NAME": shipper_contact_name,
                "SHIPPER_CONTACT_NUMBER": shipper_contact_number,
                "SHIPPER_ADDRESS": shipper_address_text,
                "CONSIGNEE": consignee if consignee != "Add a Consignee" else "",
                "CONSIGNEE_ADDRESS": consignee_address_text,
                "POL": pol,
                "POD": pod if pod != "Add a POD" else "",
                "VESSEL": vessel if vessel != "Add a Vessel" else "",
                "CARGO": cargo,
                "TECHNICAL_NAME": technical_name,
                "CLASS": cargo_class,
                "UNNO": unno,
                "SUBRISK": subrisk,
                "PACKING_GROUP": packing_group,
                "EMS": ems,
                "FLASH_POINT": flash_point,
                "MARINE_POLLUTANT": marine_pollutant,
                "LIMITED_QUANTITY": limited_quantity,
                "NATURE_OF_CARGO": nature_of_cargo,
                "MFAG_NUMBER": mfag_number,
                "OUTER_PACKAGE": outer_package,
                "INNER_PACKAGE": inner_package,
                "GROSS_WT": gross_wt,
                "NET_WT": net_wt,
                "EQUIPMENT_TYPE": equipment_type,
                "QTY_EQUIPMENT": combined_qty_equipment,
                "QUANTITY": qty,
                "CONTAINER_NUMBER": container_number,
                "SEAL_NUMBER": seal_number
            }

            if selected_template_path.lower().endswith(".docx"):
                doc = DocxTemplate(selected_template_path)
                doc.render(data)
                tmp_docx = tempfile.NamedTemporaryFile(delete=False, suffix=".docx")
                doc.save(tmp_docx.name)
                st.success("Document generated successfully!")
                st.download_button(
                    "Download Filled Word Document",
                    data=open(tmp_docx.name, "rb").read(),
                    file_name=f"{selected_template_name}_filled.docx",
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                )
            elif selected_template_path.lower().endswith((".xlsx", ".xls", ".xltx")):
                filled_xlsx_path = fill_excel_template(selected_template_path, data)
                st.success("Excel document generated successfully!")
                st.download_button(
                    "Download Filled Excel Document",
                    data=open(filled_xlsx_path, "rb").read(),
                    file_name=f"{selected_template_name}_filled.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.error("Unsupported template format. Please use .docx or .xlsx files.")

if __name__ == "__main__":
    main()