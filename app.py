import streamlit as st
import json
import tempfile
import pandas as pd
import textwrap
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io
import os

# ----------------- Google API Helper ------------------

SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/spreadsheets.readonly"
]

class GoogleDriveSheets:
    def __init__(self, service_account_info):
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".json")
        tmp.write(json.dumps(service_account_info).encode())
        tmp.close()
        self.SERVICE_ACCOUNT_FILE = tmp.name

        self.creds = service_account.Credentials.from_service_account_file(
            self.SERVICE_ACCOUNT_FILE, scopes=SCOPES
        )
        self.drive_service = build('drive', 'v3', credentials=self.creds)
        self.sheets_service = build('sheets', 'v4', credentials=self.creds)

    def download_spreadsheet_as_df(self, spreadsheet_id, sheet_name):
        result = self.sheets_service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=sheet_name
        ).execute()
        values = result.get('values', [])
        if not values:
            return pd.DataFrame()

        if sheet_name == "Equipment Type":
            df = pd.DataFrame(values, columns=["Equipment Type"])
        else:
            header = values[0]
            data_rows = values[1:]
            max_cols = len(header)
            padded_rows = []
            for row in data_rows:
                if len(row) < max_cols:
                    row += [""] * (max_cols - len(row))
                elif len(row) > max_cols:
                    row = row[:max_cols]
                padded_rows.append(row)
            df = pd.DataFrame(padded_rows, columns=header)
        return df

    def list_folder_files(self, folder_id, mime_types=None):
        query = f"'{folder_id}' in parents and trashed = false"
        if mime_types:
            mime_filter = " or ".join([f"mimeType='{mime}'" for mime in mime_types])
            query += f" and ({mime_filter})"
        files = []
        page_token = None
        while True:
            response = self.drive_service.files().list(
                q=query,
                spaces='drive',
                fields='nextPageToken, files(id, name, mimeType)',
                pageToken=page_token
            ).execute()
            files.extend(response.get('files', []))
            page_token = response.get('nextPageToken', None)
            if page_token is None:
                break
        return files

    def download_file_to_temp(self, file_id, file_name):
        request = self.drive_service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        fh.seek(0)
        suffix = os.path.splitext(file_name)[1]
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        tmp.write(fh.read())
        tmp.flush()
        tmp.close()
        return tmp.name

# ----------------- Helper functions ------------------

def convert_keys_to_template_style(orig_dict):
    key_map = {
        "technicalName": "TECHNICAL_NAME",
        "class": "CLASS",
        "unno": "UNNO",
        "subrisk": "SUBRISK",
        "packingGroup": "PACKING_GROUP",
        "ems": "EMS",
        "flashPoint": "FLASH_POINT",
        "marinePollutant": "MARINE_POLLUTANT",
        "Limited Quantity ": "LIMITED_QUANTITY",
        "natureOfCargo": "NATURE_OF_CARGO",
        "MFAG Number": "MFAG_NUMBER"
    }
    return {key_map.get(k, k).replace(" ", "_"): v for k, v in orig_dict.items()}

def format_address_by_chars(address, width=40):
    if not address:
        return ""
    return "\n".join(textwrap.wrap(address, width=width))

def fill_excel_template(template_path, data):
    wb = load_workbook(template_path)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for key, val in data.items():
                        placeholder = "{{" + key + "}}"
                        if placeholder in cell.value:
                            if key in ("SHIPPER_ADDRESS", "CONSIGNEE_ADDRESS"):
                                val = val.replace("\n", "\n")
                                cell.value = cell.value.replace(placeholder, val)
                                cell.alignment = Alignment(wrap_text=True)
                            else:
                                cell.value = cell.value.replace(placeholder, str(val))
    tmp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp_xlsx.name)
    return tmp_xlsx.name

# ----------------- Load Master Data ------------------

@st.cache_data(ttl=3600)
def load_master_data_from_drive(service_account_info, spreadsheet_id):
    gds = GoogleDriveSheets(service_account_info)
    cargo_df = gds.download_spreadsheet_as_df(spreadsheet_id, "cargo")
    equipment_df = gds.download_spreadsheet_as_df(spreadsheet_id, "Equipment Type")
    shippers_df = gds.download_spreadsheet_as_df(spreadsheet_id, "Shippers")
    consignees_df = gds.download_spreadsheet_as_df(spreadsheet_id, "Consignees")
    ports_df = gds.download_spreadsheet_as_df(spreadsheet_id, "Ports")

    vessels_df = pd.DataFrame()
    try:
        vessels_df = gds.download_spreadsheet_as_df(spreadsheet_id, "Vessels")
    except Exception:
        vessels_df = pd.DataFrame()

    cargo_data = {}
    for _, row in cargo_df.iterrows():
        cargo_name = row["Proper Shipping Name"]
        details = row.drop(labels=["Proper Shipping Name"]).to_dict()
        details = {k: ("" if pd.isna(v) else v) for k, v in details.items()}
        cargo_data[cargo_name] = details

    equipment_types = equipment_df.iloc[:, 0].dropna().tolist()
    shippers = shippers_df["Shipper"].dropna().tolist()
    shipper_contacts = shippers_df.set_index("Shipper")[["ContactName", "ContactNumber", "Shipper_Address"]].to_dict(orient="index")
    consignees = consignees_df["Consignee"].dropna().tolist()
    consignee_addresses = consignees_df.set_index("Consignee")["Consignee_Address"].to_dict()

    pol_ports = ports_df["POL"].dropna().tolist() if "POL" in ports_df.columns else []
    pod_ports = ports_df["POD"].dropna().tolist() if "POD" in ports_df.columns else []

    vessels = vessels_df["Vessel_Name"].dropna().tolist() if not vessels_df.empty else []

    return cargo_data, shippers, shipper_contacts, consignees, consignee_addresses, pol_ports, pod_ports, equipment_types, vessels

# ----------------- Load Templates ------------------

@st.cache_resource(ttl=3600)
def load_templates_from_drive(service_account_info, templates_folder_id):
    gds = GoogleDriveSheets(service_account_info)
    files = gds.list_folder_files(templates_folder_id, mime_types=[
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ])
    templates = {}
    for f in files:
        path = gds.download_file_to_temp(f["id"], f["name"])
        name = os.path.splitext(f["name"])[0]
        templates[name] = path
    return templates

# ----------------- Main App ------------------

def main():
    service_account_json_str = st.secrets["google"]["service_account_json"]
    service_account_info = json.loads(service_account_json_str)

    TEMPLATES_FOLDER_ID = "1cYa1GvGGtKxP1TPD-ACzA281LpxWZNmy"
    MASTER_SPREADSHEET_ID = "1Hj0XV7Pe4oC7Mu7_x7jNvuJzOmDT5qZtlhqSytcdDno"

    cargo_data_raw, shippers, shipper_contacts, consignees, consignee_addresses, pol_ports, pod_ports, equipment_types, vessels = load_master_data_from_drive(service_account_info, MASTER_SPREADSHEET_ID)
    cargo_data = {k: convert_keys_to_template_style(v) for k, v in cargo_data_raw.items()}

    templates = load_templates_from_drive(service_account_info, TEMPLATES_FOLDER_ID)

    st.title("Hazardous DG Form Generator")

    template_options = ["Select a Template"] + list(templates.keys())
    selected_template_name = st.selectbox("Select Template", template_options)

    if selected_template_name == "Select a Template":
        st.warning("Please select a template to proceed.")
        st.stop()
    else:
        selected_template_path = templates[selected_template_name]

    cargo = st.selectbox("Select Proper Shipping Name", list(cargo_data.keys()))
    if cargo and cargo in cargo_data:
        details = cargo_data[cargo]
    else:
        details = {k: "" for k in [
            "TECHNICAL_NAME", "CLASS", "UNNO", "SUBRISK", "PACKING_GROUP", "EMS",
            "FLASH_POINT", "MARINE_POLLUTANT", "LIMITED_QUANTITY", "NATURE_OF_CARGO", "MFAG_NUMBER"]}

    technical_name = st.text_input("Technical Name", details.get("TECHNICAL_NAME", ""))
    cargo_class = st.text_input("Class", details.get("CLASS", ""))
    unno = st.text_input("UNNO", details.get("UNNO", ""))
    subrisk = st.text_input("Subrisk", details.get("SUBRISK", ""))
    packing_group = st.text_input("Packing Group", details.get("PACKING_GROUP", ""))
    ems = st.text_input("EMS", details.get("EMS", ""))
    flash_point = st.text_input("Flash Point", details.get("FLASH_POINT", ""))

    marine_pollutant_options = ["YES", "NO"]
    marine_pollutant_default = details.get("MARINE_POLLUTANT", "").upper()
    marine_pollutant = st.selectbox(
        "Marine Pollutant", marine_pollutant_options,
        index=marine_pollutant_options.index(marine_pollutant_default) if marine_pollutant_default in marine_pollutant_options else 1)

    limited_quantity_options = ["YES", "NO", "-"]
    limited_quantity_default = details.get("LIMITED_QUANTITY", "").upper()
    limited_quantity = st.selectbox(
        "Limited Quantity", limited_quantity_options,
        index=limited_quantity_options.index(limited_quantity_default) if limited_quantity_default in limited_quantity_options else 2)

    nature_of_cargo_options = ["SOLID", "LIQUID", "GAS", "-"]
    nature_of_cargo_default = details.get("NATURE_OF_CARGO", "").upper()
    nature_of_cargo = st.selectbox(
        "Nature of Cargo", nature_of_cargo_options,
        index=nature_of_cargo_options.index(nature_of_cargo_default) if nature_of_cargo_default in nature_of_cargo_options else 3)

    mfag_number = st.text_input("MFAG Number", details.get("MFAG_NUMBER", ""))

    shipper = st.selectbox("Shipper", shippers)
    if shipper and shipper in shipper_contacts:
        contact_name = shipper_contacts[shipper]["ContactName"]
        contact_number = shipper_contacts[shipper]["ContactNumber"]
        shipper_address = shipper_contacts[shipper].get("Shipper_Address", "")
    else:
        contact_name = ""
        contact_number = ""
        shipper_address = ""

    formatted_shipper_address = format_address_by_chars(shipper_address, width=40)
    shipper_contact_name = st.text_input("Shipper Contact Name", contact_name)
    shipper_contact_number = st.text_input("Shipper Contact Number", contact_number)
    shipper_address_text = st.text_area("Shipper Address", formatted_shipper_address)

    consignee = st.selectbox("Consignee", consignees)
    if consignee and consignee in consignee_addresses:
        consignee_address = consignee_addresses[consignee]
    else:
        consignee_address = ""

    formatted_consignee_address = format_address_by_chars(consignee_address, width=40)
    consignee_address_text = st.text_area("Consignee Address", formatted_consignee_address)

    pol = st.selectbox("Port of Loading (POL)", pol_ports)
    pod = st.selectbox("Port of Discharge (POD)", pod_ports)
    vessel = st.selectbox("Vessel", vessels)

    qty = st.number_input("Quantity", min_value=1, step=1, value=1)
    equipment_type = st.selectbox("Equipment Type *", equipment_types)

    outer_package = st.text_input("Outer Package *")
    inner_package = st.text_input("Inner Package *")
    gross_wt = st.text_input("Gross Weight *")
    net_wt = st.text_input("Net Weight *")

    mandatory_fields_filled = all([
        outer_package.strip(),
        inner_package.strip(),
        gross_wt.strip(),
        net_wt.strip(),
        equipment_type != ""
    ])

    combined_qty_equipment = f"{qty}x{equipment_type}" if equipment_type != "" and qty > 0 else ""

    if st.button("Generate Document"):
        if not mandatory_fields_filled:
            st.error("Please fill all mandatory fields marked with * and select Equipment Type")
        else:
            data = {
                "SHIPPER": shipper,
                "SHIPPER_CONTACT_NAME": shipper_contact_name,
                "SHIPPER_CONTACT_NUMBER": shipper_contact_number,
                "SHIPPER_ADDRESS": shipper_address_text,
                "CONSIGNEE": consignee,
                "CONSIGNEE_ADDRESS": consignee_address_text,
                "POL": pol,
                "POD": pod,
                "VESSEL": vessel,
                "CARGO": cargo,
                "TECHNICAL_NAME": technical_name,
                "CLASS": cargo_class,
                "UNNO": unno,
                "SUBRISK": subrisk,
                "PACKING_GROUP": packing_group,
                "EMS": ems,
                "FLASH_POINT": flash_point,
                "MARINE_POLLUTANT": marine_pollutant,
                "LIMITED_QUANTITY": limited_quantity,
                "NATURE_OF_CARGO": nature_of_cargo,
                "MFAG_NUMBER": mfag_number,
                "OUTER_PACKAGE": outer_package,
                "INNER_PACKAGE": inner_package,
                "GROSS_WT": gross_wt,
                "NET_WT": net_wt,
                "EQUIPMENT_TYPE": equipment_type,
                "QTY_EQUIPMENT": combined_qty_equipment,
                "QUANTITY": qty
            }

            if selected_template_path.lower().endswith(".docx"):
                doc = DocxTemplate(selected_template_path)
                doc.render(data)
                tmp_docx = tempfile.NamedTemporaryFile(delete=False, suffix=".docx")
                doc.save(tmp_docx.name)
                st.success("Document generated successfully!")
                st.download_button(
                    "Download Filled Word Document",
                    data=open(tmp_docx.name, "rb").read(),
                    file_name=f"{selected_template_name}_filled.docx",
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                )
            elif selected_template_path.lower().endswith((".xlsx", ".xls", ".xltx")):
                filled_xlsx_path = fill_excel_template(selected_template_path, data)
                st.success("Excel document generated successfully!")
                st.download_button(
                    "Download Filled Excel Document",
                    data=open(filled_xlsx_path, "rb").read(),
                    file_name=f"{selected_template_name}_filled.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.error("Unsupported template format. Please use .docx or .xlsx files.")

if __name__ == "__main__":
    main()
